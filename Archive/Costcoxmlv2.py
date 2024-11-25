# %%

import xml.etree.ElementTree as ET
from datetime import datetime
from tkinter.filedialog import askopenfilename
import openpyxl
from openpyxl import Workbook

xml_file = "/Users/d3ops/Downloads/PO_1924323739.xml"
#xml_file = askopenfilename()

tree = ET.parse(xml_file)
root = tree.getroot()

def format_date(date_string):
    return datetime.strptime(date_string, '%Y%m%d').strftime('%d/%m/%Y')

#gather all the key data
contact_name = root.find('.//FROMBUSS')
po_number = root.find('.//PONUMBER')
inv_number = f"INV-{po_number.text}"
podate = root.find('.//BGM/PODATE')
deldate = root.find('.//DTM/REQUIREDDELIVERYDATE')
duedate = deldate.text
currency = root.find('.//CUX/CURRENCYTYPE')
shipaddress = root.find('.//HEADER/NAD/NADST')
item_no = root.find('.//DETAIL/LINGRP/PIA/IN')
item_desc = root.find('.//DETAIL/LINGRP/IMD')
item_qty = root.find('.//DETAIL/LINGRP/QTY/TOTALORDEREDQUANTITY')
item_price = root.find('.//DETAIL/LINGRP/PRI/UNITCOSTEXCLUSIVEGST')
tax_type = "GST on Income"
if currency.text == "AUD":
    acct_code = "225110723"
elif currency.text == "NZD":
    acct_code = "225110713"
else:
    acct_code = "Acct code error"

# Initialize a list to collect text content
keyaddelements = 8

addresselements = []

# Iterate through the first n elements and collect their text
for i, elem in enumerate(shipaddress.iter()):
    if i < keyaddelements:
        text = elem.text.strip() if elem.text else ''  # Use an empty string if no text is present
        addresselements.append(text)
    else:
        break

full_address = ' '.join(addresselements).lstrip()

# %%

#build a new excel workbook and populate the data
col_indexes = [1,11,12,13,14,16,17,18,20,21]

column_names = [
    "*ContactName", "EmailAddress", "POAddressLine1", "POAddressLine2", "POAddressLine3",
    "POAddressLine4", "POCity", "PORegion", "POPostalCode", "POCountry", "*InvoiceNumber", "Reference", "*InvoiceDate",
    "*DueDate", "InventoryItemCode", "*Description", "*Quantity", "*UnitAmount", "Discount", "*AccountCode",
    "*TaxType", "TrackingName1", "TrackingOption1", "TrackingName2", "TrackingOption2", "Currency", "BrandingTheme"
]

wb = Workbook()
ws = wb.active

for col_num, column_title in enumerate(column_names, start=1):
    ws.cell(row=1, column=col_num, value=column_title)

# Iterate over rows
for row in range(2,3):
    ws.cell(row=row, column=col_indexes[0], value=contact_name.text)
    ws.cell(row=row, column=col_indexes[1], value=inv_number)
    ws.cell(row=row, column=col_indexes[2], value=po_number.text)
    ws.cell(row=row, column=col_indexes[3], value=format_date(deldate.text))
    ws.cell(row=row, column=col_indexes[4], value=format_date(duedate))
    ws.cell(row=row, column=col_indexes[5], value=item_desc.text)
    ws.cell(row=row, column=col_indexes[6], value=item_qty.text)
    ws.cell(row=row, column=col_indexes[7], value=item_price.text)
    ws.cell(row=row, column=col_indexes[8], value=acct_code)
    ws.cell(row=row, column=col_indexes[9], value=tax_type)

wb.save(filename = "/Users/d3ops/Downloads/" + str(contact_name.text) + "-" + str(po_number.text) + ".xlsx")

# %% redundant data checks


# print(f"Customer: {contact_name.text}")    
# print(f"PO Number: {po_number.text}")
# print(f"PO Date: {format_date(podate.text)}")
# print(f"Delivery Date: {format_date(deldate.text)}")
# print(f"Costco Item #: {item_no.text}")
# print(f"Costco Item Description: {item_desc.text}")
# print(f"Costco Item Quantity: {item_qty.text}")
# print(f"Costco Item Price {currency.text}: {item_price.text}")

# print(f"Ship to: {full_address}")

# print(inv_number)

#key_cols = ["contact_name", "po_number", "inv_number", "del_date", "due_date","item_desc", "item_qty", "item_price", "acct_code", "tax_type"]
# %%
